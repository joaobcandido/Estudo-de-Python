import openpyxl

# criar uma planilha(book)
book = openpyxl.Workbook("vazio")

# como visualizar paginas existentes na planilha criada
print(book.sheetnames)

# como criar uma pagina 
book.create_sheet("Frutas")


# como selecionar uma pagina
frutas_page = book["Frutas"]


# adicionando valores em linha na pagina Frutas
frutas_page.append(["Frutas","Quantidade","Preco"])
frutas_page.append(["Banana","5","R$5,00"])
frutas_page.append(["Mamao","9","R$12,00"])
frutas_page.append(["Pera","8","R$15,00"])
frutas_page.append(["Macauva","88","R$15,00"])
frutas_page.append(["Pitaia","3","R$215,00"])
frutas_page.append(["Pepin","30","R$15,00"])



#salvar a planilha
book.save("Planilha de Compras.xlsx")